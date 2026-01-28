import openpyxl

df = openpyxl.load_workbook("data.xlsx")
df1 = df.active

for row in range(0, df1.max_row):
    for col in df1.iter_cols(1, df1.max_column):
        print(col[row].value)
print('\n\n\n\n\n\n\n\n\n\n') 
import xlwings as xw

# Specifying a sheet
ws = xw.Book("data.xlsx").sheets['Sheet1']

v1 = ws.range("B1:B7").value

print("Result:", v1)